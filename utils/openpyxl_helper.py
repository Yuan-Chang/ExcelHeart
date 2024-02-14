import os
import sys

from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import List
from copy import copy
from utils.Utils import get_file_name_without_extension_from_file_path
from utils.modules import Sheet
from openpyxl.utils.dataframe import dataframe_to_rows
import base64


def copy_worksheet(source_ws, target_ws, copy_value: bool = True):
    for row in source_ws.iter_rows():
        for cell in row:
            if copy_value:
                target_ws[cell.coordinate].value = cell.value

            target_ws[cell.coordinate].font = copy(cell.font)
            target_ws[cell.coordinate].border = copy(cell.border)
            target_ws[cell.coordinate].fill = copy(cell.fill)
            target_ws[cell.coordinate].number_format = cell.number_format
            target_ws[cell.coordinate].protection = copy(cell.protection)
            target_ws[cell.coordinate].alignment = copy(cell.alignment)
            target_ws[cell.coordinate].comment = cell.comment

    # Copy cell width and height
    for idx, rd in source_ws.row_dimensions.items():
        target_ws.row_dimensions[idx] = copy(rd)

    for idx, rd in source_ws.column_dimensions.items():
        target_ws.column_dimensions[idx] = copy(rd)

    for merged_cell in source_ws.merged_cells:
        target_ws.merge_cells(f"{merged_cell}")


def merge_multiple_excels_to_one_excel(input_files: List, output_file):
    output_wb = Workbook()

    for file_name in input_files:
        source_wb = load_workbook(file_name)
        number_of_sheets = len(source_wb.sheetnames)

        for sheet in source_wb:
            file_name = get_file_name_without_extension_from_file_path(file_name)
            sheet_name = f"{file_name}_{sheet.title}"

            # if only 1 sheet, use file name instead
            if number_of_sheets == 1:
                sheet_name = file_name

            output_ws = output_wb.create_sheet(sheet_name)
            copy_worksheet(sheet, output_ws)

    # Delete the default sheet
    del output_wb["Sheet"]
    output_wb.save(output_file)


# merge wb2 into wb1
def merge_two_workbook(wb1: Workbook, wb2: Workbook):
    source_wb = wb2

    for sheet in source_wb:
        sheet_name = sheet.title

        output_ws = wb1.create_sheet(sheet_name)
        copy_worksheet(sheet, output_ws)


def write_sheet_to_worksheet(wb: Workbook, sheets: List[Sheet]):
    for sheet in sheets:
        ws = wb.create_sheet(sheet.name)
        rows = dataframe_to_rows(sheet.data_frame, index=False)

        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)


def save_file_as_xlsm(workbook: Workbook, result_file, empty_macro_file):
    macro_wb = load_workbook(empty_macro_file, keep_vba=True)
    merge_two_workbook(macro_wb, workbook)
    macro_wb.save(result_file)


def delta_check(current_ws: Worksheet, previous_ws: Worksheet, delta_ws: Worksheet, start_point: List,
                header_count: int = 2):
    start_row_index = start_point[0] - 1
    start_column_index = column_index_from_string(start_point[1]) - 1

    current_header_keys = get_header_keys(start_row_index, start_column_index, current_ws, header_count=header_count)
    current_data_columns = get_data_columns(start_row_index, start_column_index, current_ws, header_count=header_count)

    previous_header_keys = get_header_keys(start_row_index, start_column_index, previous_ws, header_count=header_count)
    previous_data_columns = get_data_columns(start_row_index, start_column_index, previous_ws,
                                             header_count=header_count)

    # h1 h2 "" h3
    # h2 "" h3
    header_dict = {}
    for header_column_index, header in enumerate(previous_header_keys):
        if header != "":
            header_dict[header] = header_column_index

    # Calculate and write the delta cell value
    for header_column_index, header in enumerate(current_header_keys):

        if header in header_dict:
            # value check
            current_data_column = current_data_columns[header_column_index]
            previous_data_column = previous_data_columns[header_dict[header]]

            for data_row_index, current_cell in enumerate(current_data_column):
                previous_cell = previous_data_column[data_row_index]

                result_value = f"='{current_ws.title}'!{current_cell.coordinate}-'{previous_ws.title}'!{previous_cell.coordinate}"

                delta_ws[current_cell.coordinate].value = result_value

    # Copy the rest
    for r_index, row in enumerate(current_ws.iter_rows()):
        for c_index, cell in enumerate(row):
            if r_index < start_row_index + header_count or c_index < start_column_index:
                delta_ws[cell.coordinate].value = cell.value

    copy_worksheet(current_ws, delta_ws, copy_value=False)


def get_header_keys(start_row_index: int, start_column_index: int, ws: Worksheet, header_count: int = 2):
    header_row_range = range(start_row_index, start_row_index + header_count)
    header_column_range = range(start_column_index, ws.max_column)

    key_dict = {}
    key_list = []

    # for row_index in header_row_range:
    for column_index in header_column_range:

        key = ""
        for row_index in header_row_range:
            cell = ws[get_coordinate(row_index, column_index)]
            merged_cell_value = get_merged_cell_value(ws, cell)

            if merged_cell_value is None:
                merged_cell_value = ""

            if key == "":
                key = merged_cell_value
            else:
                key = f"{key}_{merged_cell_value}"

        # Handle duplicate key
        # if duplicate exist, we append the count to the end
        # ex. key, key_1, key_2
        if key != "":
            if key in key_dict:
                count = key_dict[key]
                key = f"{key}_{count}"
                key_dict[key] = count + 1
            else:
                key_dict[key] = 1

        key_list.append(key)
    return key_list


def get_data_columns(start_row_index: int, start_column_index: int, ws: Worksheet, header_count: int = 2):
    header_row_range = range(start_row_index + header_count, ws.max_row)
    header_column_range = range(start_column_index, ws.max_column)

    column_list = []

    # for row_index in header_row_range:
    for column_index in header_column_range:

        column = []
        for row_index in header_row_range:
            cell = ws[get_coordinate(row_index, column_index)]
            column.append(cell)

        column_list.append(column)

    return column_list


def get_coordinate(r_index, c_index):
    col = get_column_letter(c_index + 1)
    return f"{col}{r_index + 1}"


# if the cell is part of a merged cell, return the first cell value
# If not, return cell value
def get_merged_cell_value(ws: Worksheet, cell):
    for merged_cell in ws.merged_cells.ranges:
        if cell.coordinate in merged_cell:
            pair = merged_cell.left[0]
            return ws.cell(row=pair[0], column=pair[1]).value
    return cell.value


def get_main_text():
    return "CkRlYXIgQW5uaWUKCiAgICAgICBIYXBweSBWYWxlbnRpbmUncyBEYXkgIDopKSkKICAgICAgICAKICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICAgICBBbnRob255"


def get_main_font():
    return Font(size="180", name='Calibri (Body)', bold=True, italic=True)


def draw_heart(wb: Workbook, sheet_name="heart", with_text=True, size=100, x_offset=100, y_offset=25,
               text=get_main_text(), font=get_main_font()):
    intensity = 20000
    scale = size
    center_shift_x = x_offset
    center_shift_y = y_offset

    # Get x coordinate
    c_list = []
    gap = 1. / intensity
    for i in range(0, intensity + 1):
        x = i * gap
        c_list.append(x)

        if i != 0.:
            c_list.append(-x)

    # Get y coordinate
    xy_list = []
    for x in c_list:
        # Upper side of the heart
        y = get_upper_part_heart_y(x)
        xy_list.append((x, y))

        # Bottom side of the heart
        y = get_lower_part_heart_y(x)
        xy_list.append((x, y))

    # Create the sheet
    ws = wb.create_sheet(title=sheet_name)
    del wb["Sheet"]

    # Set the data list
    data_list = xy_list

    # Adjust scale to odd number so that there is a always mid-point
    if scale % 2 == 0:
        scale = scale + 1

    # As the normalized coordinate is from -1 to 1, the unit width size is 2
    # So we divide scale by 2
    scale_times = int(scale / 2)

    scaled_data_list = []

    # The center point should be the max X, Y
    max_x = -sys.maxsize * 2 - 1
    max_y = -sys.maxsize * 2 - 1
    for data in data_list:
        x = data[0] * scale_times
        y = data[1] * scale_times

        if abs(y) > max_y:
            max_y = abs(y)

        if abs(x) > max_x:
            max_x = abs(x)

        scaled_data_list.append((x, y))

    # The center point should be the max X, Y
    center_r = max_y + center_shift_y
    center_c = max_x + center_shift_x

    p_list = []
    for xy in scaled_data_list:
        x = xy[0]
        y = xy[1]

        r = int(y + center_r)
        c = int(x + center_c)
        p_list.append((c, r))

        cell = ws[get_coordinate(int(r), int(c))]
        cell.fill = get_red_fill()

    sorted_p_list = sorted(p_list, key=lambda point: point[0])
    top_center_y = int(get_upper_part_heart_y(0) * scale_times + center_r)

    max_area = -1
    max_p = (0, 0)
    for i in range(0, int(len(sorted_p_list) / 2)):
        p = sorted_p_list[i]

        heart_left_space_width = p[0] - center_shift_x
        area = (scale - 2 * heart_left_space_width) * (p[1] - top_center_y)
        if area > max_area:
            max_area = area
            max_p = p

    heart_right_space_width = max_p[0] - center_shift_x
    start_p = (max_p[0] + 2, top_center_y + 1)
    end_p = (scale + center_shift_x - heart_right_space_width - 2, max_p[1] - 1)

    if with_text:
        ws.merge_cells(start_row=start_p[1], start_column=start_p[0], end_row=end_p[1],
                       end_column=end_p[0])
        cell = ws[get_coordinate(start_p[1] - 1, start_p[0] - 1)]
        coded_string = text
        cell.value = base64.b64decode(coded_string)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell.font = font

    # Adjust cell width and height
    for idx in range(0, ws.max_column):
        ws.column_dimensions[get_column_letter(idx + 1)].width = 7

    for idx in range(0, ws.max_row):
        ws.row_dimensions[idx].height = 30

    ws.sheet_view.zoomScale = 10


def get_red_fill():
    return PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')


def get_upper_part_heart_y(x):
    return -(x * x) ** (1. / 3.) - (1 - x * x) ** (1. / 2.)


def get_lower_part_heart_y(x):
    return -(x * x) ** (1. / 3.) + (1 - x * x) ** (1. / 2.)
